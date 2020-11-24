
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import glob
import os

wb = load_workbook("medical_data_demo.xlsx")
# print(wb.active)  # 获取电子表格是否有数据
# print(wb.read_only)  # 是否是以只读方式打开
# print(wb.encoding)  # 获取表格编码
# print('', wb.properties)  # 获取电子表格属性
# print(wb.worksheets)  # 获取工作表名
# print(wb.get_sheet_names())  # 获取工作表的所有名字
# print(wb.sheetnames)  # 同上，区别是适于最新版本
# print(wb.get_sheet_by_name('粤通'))  # 通过工作表的名字，获取worksheet对象操作电子表格
# print(wb.create_sheet("test02"))  # 创建工作表，save保存
# wb.save("medical_data_demo.xlsx")
# print(wb.copy_worksheet(wb['test01']))  # 复制工作表
# wb.save("medical_data_demo.xlsx")

sheet1_test = wb['龙一']
# print(sheet1_test.title)  # 工作表的标题
# print(sheet1_test.dimensions)  # 获取表格大小
# print(sheet1_test.max_row)  # 表格最大行数s
# print(sheet1_test.min_row)  # 表格最小行数
# print(sheet1_test.max_column)  # 表格最大列数
# print(sheet1_test.min_column)  # 表格最小列数
# print(sheet1_test.rows)  # 按行获取单元格
# print(sheet1_test.columns)  # 按列获取单元格
# print(sheet1_test.freeze_panes)  # 冻结窗口
# print(sheet1_test.values)  # 按行获取表格的内容
# print(sheet1_test.iter_rows())  # 迭代器方式，按行获取所有单元格
# print(sheet1_test.iter_columns())  # 迭代器方式，按列获取所有单元格
# sheet1_test.append(['1列', '2列', '3列', '4列'])
# wb.save("medical_data_demo.xlsx")


# print(sheet1_test.unmerged_cells)  # 取消合并单元格
# print(sheet1_test['C31'].row)  # 获取行
# print(sheet1_test['C32'].row)  # 获取行
# print(sheet1_test['A1'].column)  # 获取列
# print(sheet1_test['C2'].value)  # 获取值
# sheet1_test.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)  # 合并单元格
# sheet1_test.unmerge_cells(start_row=1, start_column=2, end_row=2, end_column=2)  # 取消合并单元格
# wb.save("medical_data_demo.xlsx")

# 创建一个电子表格
# wb = Workbook()
# # ws = wb.get_sheet_by_name("Sheet")
# wb.create_sheet('test007')
# ws = wb['test007']
# ws.title = "test1124"
# ws['A1'] = 'Haden'
# ws['A2'] = 'James'
# wb.save('testCreat.xlsx')

# 运算操作
# def process_worksheet(sheet):
#     avg_column = sheet.max_column + 1  # 平均数，存放在最后一列
#     sum_column = sheet.max_column + 2  # 求和，存放在最后第二列
#     for row in sheet.iter_rows(min_row=2, min_col=2):
#         scores = [cell.value for cell in row]  # 获取一行的值
#         sum_score = sum(scores)  # 求一行的和
#         avg_score = sum_score / len(scores)  # 求一行的平均数
#         avg_cell = sheet.cell(row=row[0].row, column=avg_column)
#         sum_cell = sheet.cell(row=row[0].row, column=sum_column)
#         avg_cell.value = avg_score  # 定位到单元格，设置总分
#         sum_cell.value = sum_score  # 定位到单元格，设置平均分
#
#         # 设置对齐方式，水平是右对齐，垂直是居中
#         align = Alignment(horizontal='left', vertical='center', wrap_text=True)
#         avg_cell.alignment = align
#         sum_cell.alignment = align
#
#     # 设置平均分和总分的标题
#     sheet.cell(row=1, column=avg_column).value = '平均分'
#     sheet.cell(row=1, column=sum_column).value = '总分'
#
# def main():
#     wb = openpyxl.load_workbook('medical_data_demo.xlsx')
#     sheet = wb.get_sheet_by_name('计算')
#     process_worksheet(sheet)
#     wb.save('medical_data_demo.xlsx')
# if __name__ == '__main__':
#     main()


# 合并单元格
def merge_xlsx_files(xlsx_files):
    wb = openpyxl.load_workbook(xlsx_files[0])  # 打开第一张电子表格
    ws = wb.active  # 激活 worksheet
    ws.title = 'merged result'  # 合并结果

    for filename in xlsx_files[1:]:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active  # 激活 worksheet
        for row in sheet.iter_rows(min_row=2):  # 从第二行开启迭代
            values = [cell.value for cell in row]  # 返回一列的值，以列表类型
            ws.append(values)  # 把列表增加到新的表格里面
    return wb

def get_all_xlsx_files(path):
    """指定后缀名，获取所有跟后缀相关的文件类型，返回列表"""
    xlsx_files = glob.glob(os.path.join(path, '*.xlsx'))
    sorted(xlsx_files, key=str.lower)  # 排序
    return xlsx_files


def main():
    path = os.path.join(os.path.dirname(os.getcwd()), '临时测试', 'excels')  # 目录自行配置
    xlsx_files = get_all_xlsx_files(path)
    wb = merge_xlsx_files(xlsx_files)
    wb.save('merge_data.xlsx')  # 保存数据到硬盘


if __name__ == '__main__':
    main()





